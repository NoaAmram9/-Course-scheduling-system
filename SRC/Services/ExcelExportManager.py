import pandas as pd
from typing import List
from SRC.Models.Course import Course
from SRC.Models.Lesson import Lesson


class ExcelExportManager:
    @staticmethod
    def format_time_range(start_hour: int, end_hour: int) -> str:
        """Convert hour range to Hebrew time format"""
        if start_hour == 0 and end_hour == 0:
            return ""
        return f"{start_hour:02d}:00-{end_hour:02d}:00"
    
    @staticmethod
    def format_day_to_hebrew(day: int) -> str:
        """Convert day number to Hebrew day"""
        day_mapping = {
            1: 'א',
            2: 'ב', 
            3: 'ג',
            4: 'ד',
            5: 'ה',
            6: 'ו',
            7: 'ש'
        }
        return day_mapping.get(day, '')
    
    @staticmethod
    def format_hebrew_time(lesson: Lesson) -> str:
        """Format complete Hebrew time string with day and hours"""
        if not lesson.time:
            return ""
        
        day_hebrew = ExcelExportManager.format_day_to_hebrew(lesson.time.day)
        time_range = ExcelExportManager.format_time_range(lesson.time.start_hour, lesson.time.end_hour)

        if day_hebrew and time_range:
            return f"{day_hebrew} {time_range}"
        return ""
    
    @staticmethod
    def lesson_type_to_hebrew(lesson_type: str) -> str:
        """Convert lesson type to Hebrew"""
        type_mapping = {
            "lecture": "הרצאה",
            "exercise": "תרגיל", 
            "lab": "מעבדה",
            "departmentHours": "ש.מחלקה",
            "reinforcement": "תגבור",
            "training": "הדרכה",
            "other": "אחר",
            "unknown": "לא ידוע"
        }
        return type_mapping.get(lesson_type, lesson_type)
    
    @staticmethod
    def format_location(building: str, room: str) -> str:
        """Format building and room to location string"""
        if building and room:
            return f"{building}-{room}"
        elif building:
            return building
        elif room:
            return room
        return ""
    
    @staticmethod
    def format_semester_to_hebrew(semester: int) -> str:
        """Convert semester number to Hebrew"""
        if semester == 1:
            return "א"
        elif semester == 2:
            return "ב"
        return ""
    
    @staticmethod  
    def format_instructors(instructors: List[str]) -> str:
        """Format instructors list to comma-separated string"""
        if not instructors:
            return ""
        return ", ".join(instructors)
    
    def export_courses_to_excel(self, courses: List[Course], file_path: str) -> bool:
        """
        Export list of courses to Excel file
        
        Args:
            courses: List of Course objects to export
            file_path: Path where to save the Excel file
            
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            # Prepare data rows
            rows = []
            
            for course in courses:
                # Get all lessons from all types
                all_lessons = []
                
                # Add lectures
                for lesson in course.lectures:
                    all_lessons.append((lesson, "lecture"))
                
                # Add exercises  
                for lesson in course.exercises:
                    all_lessons.append((lesson, "exercise"))
                
                # Add labs
                for lesson in course.labs:
                    all_lessons.append((lesson, "lab"))
                
                # Add department hours
                for lesson in course.departmentHours:
                    all_lessons.append((lesson, "departmentHours"))
                
                # Add reinforcement
                for lesson in course.reinforcement:
                    all_lessons.append((lesson, "reinforcement"))
                
                # Add training
                for lesson in course.training:
                    all_lessons.append((lesson, "training"))
                
                # If no lessons, create one row with course info only
                if not all_lessons:
                    row = {
                        'שם': course.name,
                        'קוד מלא': course.code,
                        'סוג מפגש': '',
                        'מועד': '',
                        'מרצים': '',
                        'הערה': course._notes or "",
                        'חדר': '',
                        'תקופה': self.format_semester_to_hebrew(course.semester),
                        'נ"ז': 0,
                        'ש"ש': 0
                    }
                    rows.append(row)
                else:
                    # Create row for each lesson
                    for lesson, lesson_type_override in all_lessons:
                        # Use the lesson's own type if available, otherwise use the override
                        actual_lesson_type = lesson.lesson_type if hasattr(lesson, 'lesson_type') else lesson_type_override
                        
                        # Format course code with group
                        full_code = f"{course.code}-{lesson.groupCode:02d}" if lesson.groupCode > 0 else course.code
                        
                        row = {
                            'שם': course.name,
                            'קוד מלא': full_code,
                            'סוג מפגש': self.lesson_type_to_hebrew(actual_lesson_type),
                            'מועד': self.format_hebrew_time(lesson),
                            'מרצים': self.format_instructors(lesson.instructors),
                            'הערה': course._notes or "",
                            'חדר': self.format_location(lesson.building, lesson.room),
                            'תקופה': self.format_semester_to_hebrew(course.semester),
                            'נ"ז': lesson.creditPoints if hasattr(lesson, 'creditPoints') else 0,
                            'ש"ש': lesson.weeklyHours if hasattr(lesson, 'weeklyHours') else 0
                        }
                        rows.append(row)
            
            # Create DataFrame
            df = pd.DataFrame(rows)
            
            # Ensure column order matches the original format
            column_order = ['שם', 'קוד מלא', 'סוג מפגש', 'מועד', 'מרצים', 'הערה', 'חדר', 'תקופה', 'נ"ז', 'ש"ש']
            df = df.reindex(columns=column_order)
            
            # Export to Excel
            df.to_excel(file_path, index=False, engine='openpyxl')
            
            return True
            
        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            return False
    
    def export_courses_to_excel_with_formatting(self, courses: List[Course], file_path: str) -> bool:
        """
        Export courses to Excel with enhanced formatting
        
        Args:
            courses: List of Course objects to export
            file_path: Path where to save the Excel file
            
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            # First create the basic data
            if not self.export_courses_to_excel(courses, file_path):
                return False
            
            # Load the workbook for formatting
            from openpyxl import load_workbook
            wb = load_workbook(file_path)
            ws = wb.active
            
            # Format headers
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 30)  # Cap at 30 characters
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Right-to-left text direction for Hebrew
            ws.sheet_view.rightToLeft = True
            
            # Save formatted workbook
            wb.save(file_path)
            
            return True
            
        except ImportError:
            # Fallback to basic export if openpyxl formatting features aren't available
            return self.export_courses_to_excel(courses, file_path)
        except Exception as e:
            print(f"Error with formatting: {e}")
            # Try basic export as fallback
            return self.export_courses_to_excel(courses, file_path)